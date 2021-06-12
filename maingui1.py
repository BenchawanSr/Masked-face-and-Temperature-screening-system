import BC
import maskdect
import serial
from openpyxl import Workbook
import cv2
import time

#excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'Date'
ws['B1'] = 'Time'
ws['C1'] = 'ID'
ws['D1'] = 'Name'
ws['E1'] = 'Temp'
ws['F1'] = 'Mask'
ws['G1'] = 'Status'

countrec = 0

while(True):

    barcodeData,name = BC.readBarcode()
    print(barcodeData)
    print(name)
    
    
    ser = serial.Serial('COM11',9600,timeout=0,bytesize=8,parity='N',stopbits=1)
    ser.close()
    time.sleep(1)
    if not(ser.is_open):
        ser.open()
        time.sleep(1)
    
    ser2 = serial.Serial('COM10',9600,timeout=0,bytesize=8,parity='N',stopbits=1)
    ser2.close()
    time.sleep(1)
    if not(ser2.is_open):
        ser2.open()
        time.sleep(1)

    cap = cv2.VideoCapture(0)

    font = cv2.FONT_ITALIC

    #img1 = cv2.imread("bg1.jpg",cv2.IMREAD_UNCHANGED) #BGR

    xp = 66
    yp = 284
    wp = 921
    hp = 648

    cnt = 3
    temp2=0
    door=0
    while(cap.isOpened()and(door==0)):


        ret, img = cap.read()
        img = cv2.flip(img,1)
        img1 = cv2.imread("23453.jpg",cv2.IMREAD_UNCHANGED)
        img2 = cv2.resize(img,(wp,hp))
        img2,status = maskdect.demask(img2)
        #print(status)
        img1[yp:yp+img2.shape[0],xp:xp+img2.shape[1]] = img2
        dt = time.localtime()
        mm=dt[1]
        m=mm-1
        #print(mm)
        wd=dt[6]
        hr=dt[3]
        mn=dt[4]
        sd=dt[5]
        if(0<=hr<10):
            hrr='0'+str(hr)
        elif(hr>9):
            hrr=str(hr)
        if(0<=mn<10):
            mnn='0'+str(mn)
        elif(mn>9):
             mnn=str(mn)
        if(0<=sd<10):
            sdd='0'+str(sd)
        elif(sd>9):
            sdd=str(sd)
        nameday=["MONDAY","TUESDAY","WEDNESDAY","THURSDAY","FRIDAY","SATURDAY"]
        namemon=["JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE","JULY","AUGUST","SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"]
        weekday=nameday[wd]
        mont=namemon[m]
        DD = (str(weekday)+' '+str(dt[2])+' '+str(mont)+' '+str(dt[0]))
        dd=(str(dt[2])+'/'+str(dt[1])+'/'+str(dt[0]))
        d=(str(dt[2])+'-'+str(dt[1])+'-'+str(dt[0]))
        TT = (str(hrr)+':'+str(mnn)+':'+str(sdd))
        t=(str(hrr)+'-'+str(mnn)+'-'+str(sdd))
        #print(dt)
        cv2.putText(img1,DD,(1225,320),font,1,(193,196,188),4)
        cv2.putText(img1,TT,(1175,435),font,4,(193,196,188),10)
        
        xp1 = 475
        yp1 = 860
        wp1 = 100
        hp1 = 100

        img3 = cv2.imread("woman.png",cv2.IMREAD_UNCHANGED) #BGRA
        img3 = cv2.resize(img3,(wp1,hp1))
        channels = cv2.split(img3)
        mask = channels[3]

        mask_inv = cv2.bitwise_not(mask)
        #cv2.imshow("mask",mask)
        #cv2.imshow("mask_inv",mask_inv)

        roi = img1[yp1:yp1+hp1,xp1:xp1+wp1]
        #cv2.imshow("roi",roi)

        img1_bg = cv2.bitwise_or(roi,roi,mask=mask_inv)
        img2_fg = cv2.bitwise_and(img3[:,:,:3],img3[:,:,:3],mask=mask)
        #cv2.imshow("img1_bg",img1_bg)
        #cv2.imshow("img2_fg",img2_fg)

        dst = cv2.add(img1_bg,img2_fg)

        img1[yp1:yp1+hp1,xp1:xp1+wp1] = dst
        #temp=0
        #while temp==0:

        
        print("***********************")
        if not(ser.is_open):
            ser.open()
        for temp in ser.read():
            if cnt%6 == 0:
                err =((temp-104)*0.9)+68
                print("err = ",err)
                temp = round(temp-err,1)
                print("cnt = ", cnt)
                print("temp = ", temp)
                if temp > 34 or temp < 40:
                    temp2 = temp
                else:
                    pass
            cnt=cnt+1
            
        
        
        cv2.putText(img1,str(temp2),(1175,935),font,3,(42,55,23),7)
        cv2.putText(img1,name,(1175,775),font,3,(42,55,23),7)
        cv2.putText(img1,barcodeData,(1175,600),font,3,(42,55,23),7)
        #cv2.imshow("img1+img3",img1)
        #cv2.imshow("img2",img2)
        if (temp2<37.5)and(temp2>34)and(status=="M"):
            #print("PASS")
            xp4 = 375
            yp4 = 860
            wp4 = 100
            hp4 = 100

            img4 = cv2.imread("777.png",cv2.IMREAD_UNCHANGED) #BGRA
            img4 = cv2.resize(img4,(wp4,hp4))
            channels = cv2.split(img4)
            mask = channels[3]

            mask_inv = cv2.bitwise_not(mask)
              #cv2.imshow("mask",mask)
              #cv2.imshow("mask_inv",mask_inv)

            roi = img1[yp4:yp4+hp4,xp4:xp4+wp4]
              #cv2.imshow("roi",roi)

            img1_bg = cv2.bitwise_or(roi,roi,mask=mask_inv)
            img2_fg = cv2.bitwise_and(img4[:,:,:3],img4[:,:,:3],mask=mask)
              #cv2.imshow("img1_bg",img1_bg)
              #cv2.imshow("img2_fg",img2_fg)

            dst = cv2.add(img1_bg,img2_fg)
            #cv2.imshow("dst",dst)

            img1[yp4:yp4+hp4,xp4:xp4+wp4] = dst
            
        else:
            xp5 = 575
            yp5 = 860
            wp5 = 100
            hp5 = 100

            img5 = cv2.imread("666.png",cv2.IMREAD_UNCHANGED) #BGRA
            img5 = cv2.resize(img5,(wp5,hp5))
            channels = cv2.split(img5)
            mask = channels[3]
            

            mask_inv = cv2.bitwise_not(mask)
              #cv2.imshow("mask",mask)
              #cv2.imshow("mask_inv",mask_inv)

            roi = img1[yp5:yp5+hp5,xp5:xp5+wp5]
              #cv2.imshow("roi",roi)

            img1_bg = cv2.bitwise_or(roi,roi,mask=mask_inv)
            img2_fg = cv2.bitwise_and(img5[:,:,:3],img5[:,:,:3],mask=mask)
              #cv2.imshow("img1_bg",img1_bg)
              #cv2.imshow("img2_fg",img2_fg)

            dst = cv2.add(img1_bg,img2_fg)
              #cv2.imshow("dst",dst)

            img1[yp5:yp5+hp5,xp5:xp5+wp5] = dst
        cv2.imshow("img1+img3",img1)

        if (temp2<37.5)and(temp2>34)and(status=="M"):
            try:
                ser2.write(str.encode("O"))
                door=1
                #time.sleep(3)
                cv2.imwrite(barcodeData+"_"+d+"_"+t+".jpg",img1)
                allstatus = "-PASS-"
                cv2.putText(img1,allstatus,(255,405),font,5,(0,255,0),7)
                cv2.imshow("img1+img3",img1)
                print("Hello")
                
                countrec=countrec+1
                ws['A'+str(countrec+1)] = dd
                ws['B'+str(countrec+1)] = TT
                ws['C'+str(countrec+1)] = barcodeData
                ws['D'+str(countrec+1)] = name
                ws['E'+str(countrec+1)] = str(temp2)
                ws['F'+str(countrec+1)] = status
                ws['G'+str(countrec+1)] = allstatus
                wb.save(d+".xlsx")
                status="NM"
                temp2=0
                ser2.close()
                
                ser.close()
            except:
                print("Can not send")
            print("Open")
            cv2.waitKey(5000)
        
        if (temp2>37.5)and(status=="M"):
            try:
                ser2.close()
                door=1
                #time.sleep(3)
                cv2.imwrite(barcodeData+"_"+d+"_"+t+".jpg",img1)
                allstatus = "-FAIL-"
                cv2.putText(img1,allstatus,(255,405),font,5,(0,0,255),7)
                
                cv2.imshow("img1+img3",img1)
                print("Hello")
                
                countrec=countrec+1
                ws['A'+str(countrec+1)] = dd
                ws['B'+str(countrec+1)] = TT
                ws['C'+str(countrec+1)] = barcodeData
                ws['D'+str(countrec+1)] = name
                ws['E'+str(countrec+1)] = str(temp2)
                ws['F'+str(countrec+1)] = status
                ws['G'+str(countrec+1)] = allstatus
                wb.save(d+".xlsx")
                status="NM"
                temp2=0
                ser2.close()
                ser.close()

            except:
                print("Can not send")
            print("Open")
            cv2.waitKey(5000)
        if (temp2>37.5)and(status=="NM"):
            try:
                ser2.close()
                door=1
                #time.sleep(3)
                cv2.imwrite(barcodeData+"_"+d+"_"+t+".jpg",img1)
                allstatus = "-FAIL-"
                cv2.putText(img1,allstatus,(255,405),font,5,(0,0,255),7)
                
                cv2.imshow("img1+img3",img1)
                print("Hello")
                
                countrec=countrec+1
                ws['A'+str(countrec+1)] = dd
                ws['B'+str(countrec+1)] = TT
                ws['C'+str(countrec+1)] = barcodeData
                ws['D'+str(countrec+1)] = name
                ws['E'+str(countrec+1)] = str(temp2)
                ws['F'+str(countrec+1)] = status
                ws['G'+str(countrec+1)] = allstatus
                wb.save(d+".xlsx")
                status="NM"
                temp2=0
                ser2.close()
                ser.close()

            except:
                print("Can not send")
            print("Open")
            cv2.waitKey(5000)
        if (temp2<37.5)and(temp2>34)and(status=="NM"):
            try:
                ser2.close()
                door=1
                #time.sleep(3)
                cv2.imwrite(barcodeData+"_"+d+"_"+t+".jpg",img1)
                allstatus = "-FAIL-"
                cv2.putText(img1,allstatus,(255,405),font,5,(0,0,255),7)
                
                cv2.imshow("img1+img3",img1)
                print("Hello")
                
                countrec=countrec+1
                ws['A'+str(countrec+1)] = dd
                ws['B'+str(countrec+1)] = TT
                ws['C'+str(countrec+1)] = barcodeData
                ws['D'+str(countrec+1)] = name
                ws['E'+str(countrec+1)] = str(temp2)
                ws['F'+str(countrec+1)] = status
                ws['G'+str(countrec+1)] = allstatus
                wb.save(d+".xlsx")
                status="NM"
                temp2=0
                ser2.close()
                ser.close()

            except:
                print("Can not send")
            print("Open")
            cv2.waitKey(5000)
            
        status="NM"
        #cv2.imshow("img2",img2)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()


